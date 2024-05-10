import socket
import speedtest
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

def get_local_ip():
    try:
        # Mendapatkan alamat IP lokal
        local_ip = socket.gethostbyname(socket.gethostname())
        return local_ip
    except:
        return "Unknown"

def run_speedtest():
    st = speedtest.Speedtest()
    try:
        # Melakukan pengukuran speedtest
        download_speed = st.download() / 10**6  # Mengukur kecepatan download dalam Mbps
        upload_speed = st.upload() / 10**6  # Mengukur kecepatan upload dalam Mbps
        download_speed = round(download_speed, 3)
        upload_speed = round(upload_speed, 3)
        return download_speed, upload_speed
    except Exception as e:
        print(f"Error: {e}")
        return None, None

def save_to_excel(all_results):
    # Membuat DataFrame dari hasil pengukuran
    df = pd.DataFrame(all_results)

    # Menghitung rata-rata download speed dan upload speed
    avg_download_speed = df['Download Speed (Mbps)'].mean()
    avg_upload_speed = df['Upload Speed (Mbps)'].mean()

    # Mendapatkan tanggal dan waktu saat ini
    now = datetime.now()
    timestamp_str = now.strftime("%Y%m%d_%H%M%S")
    
    # Menyimpan DataFrame ke dalam file Excel dengan nama yang mencerminkan tanggal dan waktu pengukuran
    file_name = f"speedtest_{timestamp_str}.xlsx"

    # Menyimpan DataFrame ke dalam file Excel dengan format yang lebih rapi
    df.to_excel(file_name, index=False)

    # Mengatur lebar kolom sesuai isi yang didapat
    wb = load_workbook(file_name)
    ws = wb.active
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) * 1.2))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # Menulis rata-rata hanya sekali
    avg_cell = len(all_results) + 2
    ws.cell(row=avg_cell, column=len(df.columns) + 1, value="Average Download Speed (Mbps)").font = Font(bold=True)
    ws.cell(row=avg_cell + 1, column=len(df.columns) + 1, value=avg_download_speed)
    ws.cell(row=avg_cell, column=len(df.columns) + 2, value="Average Upload Speed (Mbps)").font = Font(bold=True)
    ws.cell(row=avg_cell + 1, column=len(df.columns) + 2, value=avg_upload_speed)

    # Mengatur lebar kolom untuk kolom "Average"
    avg_col = chr(ord('A') + len(df.columns))
    avg_width = max(len("Average Download Speed (Mbps)"), len("Average Upload Speed (Mbps)")) * 1.2
    ws.column_dimensions[avg_col].width = avg_width

    wb.save(file_name)

def main():
    source_host = get_local_ip()  # Mendapatkan alamat IP lokal
    all_results = []
    print("Memulai Pengukuran Speedtest:")
    start_time = time.time() # Waktu mulai pengukuran
    while (time.time() - start_time) < 30: # Program berjalan selama 30 detik
        download_speed, upload_speed = run_speedtest() # Melakukan pengukuran speedtest
        if download_speed and upload_speed:
            # Menyimpan hasil pengukuran dalam list
            all_results.append({
                "Source Host": source_host,
                "Download Speed (Mbps)": download_speed,
                "Upload Speed (Mbps)": upload_speed,
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
        time.sleep(15) # Menunda eksekusi program selama 15 detik
    # Simpan hasil ke Excel
    save_to_excel(all_results)
    print("Pengukuran Speedtest selesai")

if __name__ == "__main__":
    main()
