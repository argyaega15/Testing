import time
import socket
import pandas as pd
from scapy.all import *
from datetime import datetime
from openpyxl import load_workbook

def get_local_ip():
    # Mendapatkan alamat IP lokal
    try:
        local_ip = socket.gethostbyname(socket.gethostname())
        return local_ip
    except:
        return "Unknown"

def traceroute(destination):
    ttl = 1
    max_hops = 15
    port = 33434
    timeout = 1
    trace_data = []
    total_time = 0  # Total time for all successful hops
    hop_count = 0  # Number of successful hops

    while ttl <= max_hops:
        packet = IP(dst=destination, ttl=ttl) / ICMP()
        start_time = time.time()
        reply = sr1(packet, verbose=False, timeout=timeout)

        if reply is not None:
            mac_address = None
            if reply.haslayer(ARP):
                mac_address = reply[ARP].hwsrc
            elif reply.haslayer(Ether):
                mac_address = reply[Ether].src

        if reply is None:
            # Hop tidak dapat dicapai, catat sebagai "*"
            trace_data.append({
                "hop": ttl,
                "ip_hope": "*",
                "time": None  # Ganti "*" dengan None
            })
        elif reply.type == 3:
            # Hop mencapai tujuan akhir
            trace_data.append({
                "hop": ttl,
                "ip_hope": f"Tujuan dicapai ({reply.src})",
                "time": (time.time() - start_time) * 1000
            })
            total_time += (time.time() - start_time) * 1000
            hop_count += 1
            break
        else:
            # Hop tercapai dengan alamat IP dan waktu yang dibutuhkan
            trace_data.append({
                "hop": ttl,
                "ip_hope": str(reply.src),
                "time (ms)": (time.time() - start_time) * 1000
            })
            total_time += (time.time() - start_time) * 1000
            hop_count += 1

        ttl += 1

    # Calculate average time for successful hops
    average_time = total_time / hop_count if hop_count > 0 else None

    return trace_data, average_time

def save_to_excel(all_results, overall_average_time):
    # Membuat DataFrame dari hasil traceroute
    df = pd.DataFrame(all_results)

    # Mendapatkan tanggal dan waktu saat ini
    now = datetime.now()
    tanggal_jam = now.strftime("%Y%m%d_%H%M%S")

    # Menyimpan DataFrame ke dalam file Excel dengan nama yang berisi tanggal dan waktu
    nama_file_excel = f"traceroute_{tanggal_jam}.xlsx"
    df.to_excel(nama_file_excel, index=False)
    print(f"Data traceroute telah disimpan ke dalam file: {nama_file_excel}")

    # Mengatur ukuran sel sesuai dengan isi yang didapat
    wb = load_workbook(nama_file_excel)
    ws = wb.active
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) * 1.2))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # Menambahkan rata-rata waktu keseluruhan ke Excel
    last_row = len(all_results) + 2
    ws[f"D{last_row}"] = "Rata-rata Waktu Keseluruhan"
    ws[f"E{last_row}"] = (overall_average_time)

    wb.save(nama_file_excel)

def main():
    destination = "google.com"  # Ganti dengan tujuan traceroute Anda
    source_host = get_local_ip()  # Mendapatkan alamat IP lokal
    all_results = []
    average_times = []

    print("Pengukuran Traceroute dimulai")

    # Waktu mulai pengukuran
    start_time = time.time()

    # mengatur bearapa lama program berjalan
    while time.time() - start_time < 30:
        result, avg_time = traceroute(destination)
        all_results.extend(result)
        if avg_time:
            average_times.append(avg_time)
        time.sleep(2)  # Menunggu 2 detik sebelum melakukan traceroute berikutnya

    # Calculate overall average time
    overall_average_time = sum(average_times) / len(average_times) if average_times else None

    # Simpan hasil ke Excel
    save_to_excel(all_results, overall_average_time)

if __name__ == "__main__":
    main()
